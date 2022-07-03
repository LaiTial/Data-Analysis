"""
기사명 : 삼성전자
1페이지 10개 기사 제목, 본문 크롤링(select)
select 선택자오 매칭되는 모든 태그 객체를 리스트로 반환

그 후 가져온 data 엑셀에 저장하기
제목, 본문식으로!
"""

import requests
from bs4 import BeautifulSoup
import openpyxl

response = requests.get("https://search.naver.com/search.naver?where=news&query=%EC%82%BC%EC%84%B1%EC%A0%84%EC%9E%90&sm=tab_tmr&nso=so:r,p:all,a:all&sort=0")
html = response.text
soup = BeautifulSoup(html, 'html.parser')

news = {}

# select 선택자와 매칭되는 모든 태그 객체를 리스트로 반환
gettitle = soup.select('.news_tit') #제목
contents = soup.select('.dsc_txt_wrap') # 기사 본문
# 여러 클래스를 선택할거면 .class.class로 하면 된다.
# 예) .episode-box.mw-work-detail

for title, content in zip(gettitle, contents):
    news[title.text] = content.text
    
wb = openpyxl.Workbook()
ws = wb.create_sheet("news")
    
# 데이터 추가
ws['A1'] = 'NewsNum'
ws['B1'] = 'Title'
ws['C1'] = 'Contents'

i = 2

for title, content in news.items():
    ws['A'+str(i)] = i-1
    ws['B'+str(i)] = title
    ws['C'+str(i)] = content
    
    i += 1

wb.save("news_data.xlsx")
