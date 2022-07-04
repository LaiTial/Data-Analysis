"""
chrome driver
쇼핑 crawling

4단계 : 엑셀 저장
상품명, 가격, 상세페이지 링크
"""
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys #드라이버가 특정 키를 입력하게 만들때 사용
import time
import openpyxl

def makeXlsx(name, tagData):
    # 엑셀 시트 open
    wb = openpyxl.Workbook() #기존 엑셀파일 open
    ws = wb.create_sheet(name)
    ws.protection.disable()
    
    # 데이터 추가
    ws['A1'] = 'Num'
    ws['B1'], ws['C1'], ws['D1'] = tagData
    
    return wb, ws

targetsite = "https://search.shopping.naver.com/search/all?query="
tagData = ['product', 'price', 'link']
keyword = input("keyword? ")

wb, ws = makeXlsx(keyword, tagData)

driver = webdriver.Chrome('../chromedriver.exe')

# get() 메소드로 가상 크롬에 크롤링할 타겟 사이트를 띄운다.
driver.get(targetsite+keyword)
driver.maximize_window() #화면을 최대화
driver.implicitly_wait(20) #드라이버 구동 후 n초 동안 기다린다.

# 스크롤 내린다
driver.find_element(By.CSS_SELECTOR, 'body').send_keys(Keys.END) # 맨 밑까지 스크롤을 내린다.
time.sleep(1) #잠깐의 텀을 둔다

# 컨테이너
infos = driver.find_elements(By.CSS_SELECTOR, "li.basicList_item__2XT81")

for i, info in enumerate(infos):
    productN = info.find_element(By.CSS_SELECTOR, '.basicList_link__1MaTN').text.strip() #상품명
    link = info.find_element(By.CSS_SELECTOR, '.basicList_link__1MaTN').get_attribute('href') #상세페이지 링크
    price = info.find_element(By.CSS_SELECTOR, '.price_num__2WUXn').text.strip() #가격
    
    ws.append([i, productN, price, link])
    
wb.save("shopping.xlsx")

    
