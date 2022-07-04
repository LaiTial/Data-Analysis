"""
[네이버 지식인 크롤링]
https://kin.naver.com/

특정 키워드 검색 후,
*5단계 - 크롤링한 데이터 엑셀 추가
"""

import requests
from bs4 import BeautifulSoup
import openpyxl

keyword = input("keyword? ")
pageNum = int(input("How much page? "))

targetSite="https://kin.naver.com/search/list.naver?query=" + keyword + "&page="
nameTag = ['Question', 'Date', 'Contents', 'Link']
knowIn = {}

# 문제 번호
num = 1

for i in range(1, pageNum+1):
    response = requests.get(targetSite+str(i))
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')
    
    # 지식인 list를 가져온다
    Q = soup.select('.basic1 > li')
    
    # 제목과 링크
    for lists in Q:
        
        #for question in Qs:
        title = lists.select_one('a').text #제목
        link = lists.select_one('a').attrs['href'] #링크
        
        date = lists.select_one('.txt_inline').text # 날짜
        content = lists.select_one('.txt_inline+dd').text # 내용 #dd:nth-of-type(2) dd:nth-of-child(3) 가능
    
        tempL = [title, date, content, link]
        knowIn["Q"+str(num)] = tempL
        
        num += 1
        
# xlsx 시트 open
wb = openpyxl.Workbook()
ws = wb.create_sheet(keyword+" save")

# 데이터 추가
ws['A1'] = 'Num'
ws['B1'], ws['C1'], ws['D1'], ws['E1'] = nameTag

i = 2

for title, contents in knowIn.items():
    ws['A'+str(i)] = i-1
    ws['B'+str(i)], ws['C'+str(i)], ws['D'+str(i)], ws['E'+str(i)] = contents
    
    i += 1
    
wb.save("지식in.xlsx")
