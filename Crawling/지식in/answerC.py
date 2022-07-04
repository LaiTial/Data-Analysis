"""
[네이버 지식인 크롤링]
https://kin.naver.com/
특정 키워드 검색 후,
*6단계 - 상세 페이지 크롤링 (각 질문의 상세페이지 답변)

상세페이지 질문명, 질문내용.
그리고 답변은 최대 3개까지 크롤링
"""

import requests
from bs4 import BeautifulSoup
import openpyxl

keyword = '향수'#input("keyword? ")
pageNum = 1#int(input("How much page? "))

targetSite="https://kin.naver.com/search/list.naver?query=" + keyword + "&page="
nameTag = ['Question', 'Date', 'Contents', 'Link']
knowIn = {}

# 문제 번호
num = 1

for i in range(1, pageNum+1):
    response = requests.get(targetSite+str(i))
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')
    
    # 지식인 list를 가져온다
    Q = soup.select('.basic1 > li')
    
    # 제목과 링크
    for lists in Q:
        
        #for question in Qs:
        link = lists.select_one('a').attrs['href'] #링크
        
        date = lists.select_one('.txt_inline').text # 날짜
        
        #상세 페이지로 가기
        response = requests.get(link)
        html = response.text
        soup = BeautifulSoup(html, 'html.parser')
        
        title = soup.select_one('.c-heading__title-inner').text.strip() #제목
        
        #예외 처리
        try:
            content = soup.select_one('.c-heading__content').text.strip() # 내용 
        except:
            # 본문이 없는경우
            content=""
        
        answerList = []
        
        try:
            # 질문의 답변을 가져온다.
            named = soup.select('.c-heading-answer__title > .title')
            answers = soup.select('.se-main-container')
            
            tryNum = 0
            for name, answer in zip(named, answers):
                #temp = []
                #temp = [name.text.strip().replace('eXpert\n\n\n\n\n\n', ''), answer.text.strip()]
                answerList.append([name.text.strip().replace('eXpert\n\n\n\n\n\n', ''), answer.text.strip()])
                
                tryNum += 1
                
                if tryNum >= 3:
                    break
        except:
            #답변이 없는 경우
            answerList = []
    
        tempL = [title, date, content, link, answerList]
        knowIn["Q"+str(num)] = tempL
        
        num += 1
        
# xlsx 시트 open
wb = openpyxl.Workbook()
ws = wb.create_sheet("지식in save")

# 데이터 추가
ws['A1'] = 'Num'
ws['B1'], ws['C1'], ws['D1'], ws['E1'] = nameTag[:4]
ws['F1'], ws['G1'] = 'Name1', 'Answer1'
ws['H1'], ws['I1'] = 'Name2', 'Answer2'
ws['J1'], ws['K1'] = 'Name3', 'Answer3'

alpha = [['F', 'G'], ['H','I'],['J','K']]

i = 2
aNum = 0
for title, contents in knowIn.items():
    ws['A'+str(i)] = i-1
    ws['B'+str(i)], ws['C'+str(i)], ws['D'+str(i)], ws['E'+str(i)] = contents[:4]
    
    for ans in contents[4]:
        name, answer = ans
        
        ws[alpha[aNum][0]+str(i)] = name
        ws[alpha[aNum][1]+str(i)] = answer
        
        aNum += 1
    aNum = 0
    
    i += 1
    
wb.save("지식in.xlsx")