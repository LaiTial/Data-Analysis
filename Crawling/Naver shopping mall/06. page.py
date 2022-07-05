"""
chrome driver
쇼핑 crawling

6단계 : 사용자로부터 크롤링할 페이지 개수를 입력받아 크롤링
상품명, 가격, 상세페이지 링크
"""
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys #드라이버가 특정 키를 입력하게 만들때 사용
import time
from bs4 import BeautifulSoup
import openpyxl

def makeXlsx(name, tagData):
    # 엑셀 시트 open
    wb = openpyxl.Workbook() #기존 엑셀파일 open
    ws = wb.create_sheet(name)
    ws.protection.disable()
    
    # 데이터 추가
    ws['A1'] = 'Num'
    ws['B1'], ws['C1'], ws['D1'], ws['E1'] = tagData
    
    return wb, ws

indexNum = 1
tagData = ['product', 'price', 'img', 'link']
keyword = input("keyword? ")
pageNum = input("How many page? ")

wb, ws = makeXlsx(keyword, tagData)

driver = webdriver.Chrome('../chromedriver.exe')

for pageN in range(1, pageNum+1):
    
    targetsite = f"https://search.shopping.naver.com/search/all?pagingIndex={pageN}&query={keyword}"

    # get() 메소드로 가상 크롬에 크롤링할 타겟 사이트를 띄운다.
    driver.get(targetsite)
    driver.maximize_window() #화면을 최대화
    driver.implicitly_wait(20) #드라이버 구동 후 n초 동안 기다린다.
    preData = 0
    
    # 스크롤 내린다
    while True:
    
        driver.find_element(By.CSS_SELECTOR, 'body').send_keys(Keys.END) # 맨 밑까지 스크롤을 내린다.
        time.sleep(1) #잠깐의 텀을 둔다
        
        nowData = len(driver.find_elements(By.CSS_SELECTOR, 'li'))
    
        if(nowData == preData):
            break;
        else:
            preData = nowData
    
    time.sleep(1)
    
    # 페이지 소스를 얻어온다.
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    
    # 컨테이너
    contentBox = soup.select('.basicList_item__2XT81')
    
    print("\n")
    
    for info in contentBox:
        
        product = info.select_one('.basicList_link__1MaTN') #a 태그 가져온다
        
        productN = product.text.strip() #상품명
        link = product.attrs['href'] #상세페이지 링크
        price = info.select_one('.price_num__2WUXn').text.strip() #가격
        
        try:
            img = info.select_one('.thumbnail_thumb__3Agq6 > img').attrs['src'] #이미지 경로 가져온다
        except:
            continue;
        
        ws.append([indexNum, productN, price, img, link])
        
        indexNum +=1
        
wb.save("shopping.xlsx")
