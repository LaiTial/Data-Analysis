"""
chrome driver
youtube crawling

바로 보이는 제목, 조회수, 영상 업로드한 날짜를 crawling
실시간 생방송일 경우 skip.
"""
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys #드라이버가 특정 키를 입력하게 만들때 사용
import openpyxl

# 엑셀 시트 open
wb = openpyxl.Workbook()
ws = wb.create_sheet("풀의 신")

# 데이터 추가
ws['A1'] = 'Num'
ws['B1'], ws['C1'], ws['D1'] = '제목', '조회수', '날짜'

i = 2

driver = webdriver.Chrome('../chromedriver.exe')

# get() 메소드로 가상 크롬에 크롤링할 타겟 사이트를 띄운다.
driver.get("https://www.youtube.com/results?search_query=%ED%92%80%EC%9D%98+%EC%8B%A0")
driver.maximize_window() #화면을 최대화
driver.implicitly_wait(5) #드라이버 구동 후 n초 동안 기다린다.

# 처음 접속했을 때 영상의 날짜, 조회수, 제목 크롤링

# 컨테이너
infos = driver.find_elements(By.CSS_SELECTOR, "div.text-wrapper")

for info in infos:
    videoC = info.find_element(By.CSS_SELECTOR, '.style-scope.ytd-video-renderer')
    
    video = videoC.text.split('\n')
    
    if(len(video) != 3):
        continue;
    video.insert(0, i-1)
    ws.append(video)
    
    i += 1
    
wb.save("youtube.xlsx")