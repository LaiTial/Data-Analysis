"""
chrome driver
youtube crawling

사용자로부터 키워드를 입력받아 검색.
스크롤을 n번 내리고 제목, 조회수, 영상 업로드한 날짜를 crawling
실시간 생방송일 경우 skip.

데이터 전처리.
조회수 데이터를 숫자로 변환

기존 엑셀 파일에 저장! 새로운 시트로!
"""
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys #드라이버가 특정 키를 입력하게 만들때 사용
import openpyxl
import time
import chData as chD

def makeXlsx(filename, name, tagData):
    # 엑셀 시트 open
    wb = openpyxl.load_workbook(filename) #기존 엑셀파일 open
    ws = wb.create_sheet(name)
    ws.protection.disable()
    
    # 데이터 추가
    ws['A1'] = 'Num'
    ws['B1'], ws['C1'], ws['D1'] = tagData
    
    return wb, ws
    
filename = 'youtube.xlsx'
tagData = ['제목', '조회수', '날짜']

keyword = input("keyword? ")

wb, ws = makeXlsx(filename, keyword, tagData)

driver = webdriver.Chrome('../chromedriver.exe')

# get() 메소드로 가상 크롬에 크롤링할 타겟 사이트를 띄운다.
driver.get("https://www.youtube.com/results?search_query="+keyword)
driver.maximize_window() #화면을 최대화
driver.implicitly_wait(15) #드라이버 구동 후 n초 동안 기다린다.

# 스크롤을 n번 내리고 크롤링
for i in range(5):
    driver.find_element(By.CSS_SELECTOR, 'body').send_keys(Keys.END) # 맨 밑까지 스크롤을 내린다.
    time.sleep(1) #잠깐의 텀을 둔다

# 영상의 날짜, 조회수, 제목 크롤링

# 컨테이너
infos = driver.find_elements(By.CSS_SELECTOR, "div.text-wrapper")

for i, info in enumerate(infos):
    videoC = info.find_element(By.CSS_SELECTOR, '.style-scope.ytd-video-renderer')
    #metadata-line>span:nth-child(n)으로 n번째 선택 가능.
    
    video = videoC.text.split('\n')
    
    # 생방송은 제외!
    if(len(video) != 3):
        continue;
        
    # 조회수 데이터 자료형 변경
    video[1] = chD.text_to_num(video[1])
    
    # 번호 추가
    video.insert(0, i+1)
    
    # 데이터 엑셀에 저장
    ws.append(video)
    
    
wb.save(filename)