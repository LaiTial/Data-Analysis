"""
chrome driver
youtube crawling

스크롤을 n번 내리고
제목, 조회수, 영상 업로드한 날짜를 crawling
실시간 생방송일 경우 skip.

새로 엑셀 파일을 만들어서 저장
"""
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys #드라이버가 특정 키를 입력하게 만들때 사용
import openpyxl
import time

# 엑셀 시트 open
wb = openpyxl.Workbook() #새로 엑셀파일 생성
ws = wb.create_sheet("클레")
ws.protection.disable() #시트 보호기능 해제

# 데이터 추가
ws['A1'] = 'Num'
ws['B1'], ws['C1'], ws['D1'] = '제목', '조회수', '날짜'

driver = webdriver.Chrome('../chromedriver.exe')

# get() 메소드로 가상 크롬에 크롤링할 타겟 사이트를 띄운다.
driver.get("https://www.youtube.com/results?search_query=%ED%81%B4%EB%A0%88")
driver.maximize_window() #화면을 최대화
driver.implicitly_wait(15) #드라이버 구동 후 n초 동안 기다린다.

# 스크롤을 n번 내리고 크롤링
for i in range(5):
    driver.find_element(By.CSS_SELECTOR, 'body').send_keys(Keys.END) # 맨 밑까지 스크롤을 내린다.
    time.sleep(1) #잠깐의 텀을 둔다

# 영상의 날짜, 조회수, 제목 크롤링

# 컨테이너
infos = driver.find_elements(By.CSS_SELECTOR, "div.text-wrapper")

for i, info in enumerate(infos):
    videoC = info.find_element(By.CSS_SELECTOR, '.style-scope.ytd-video-renderer')
    #metadata-line>span:nth-child(n)으로 n번째 선택 가능.
    
    video = videoC.text.split('\n')
    
    # 생방송은 제외!
    if(len(video) != 3):
        continue;
    video.insert(0, i+1)
    ws.append(video)
    
    
wb.save("youtube.xlsx")
